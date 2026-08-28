#!/usr/bin/env python3
"""
migrate_workbook.py -- load CPDE Excel workbook data into the v2 Postgres schema.

Usage:
    python migrate_workbook.py --workbook path\\to\\CDA_Pipeline_v2_15.xlsm --dry-run
    python migrate_workbook.py --workbook path\\to\\CDA_Pipeline_v2_15.xlsm \
        --dsn "postgresql://cpde:localdev@localhost:5433/cpde" \
        --client-code COLLINS --client-name "Collins Aerospace" \
        --business "Mission Systems" --business-unit "MS BU"

--dry-run parses and validates WITHOUT touching the database. Run it first.

Requires: openpyxl, psycopg[binary]
    pip install openpyxl "psycopg[binary]"
"""

import argparse
import sys
from collections import OrderedDict

import openpyxl

# ---------------------------------------------------------------------
# Sheet layout (workbook v2.15)
# ---------------------------------------------------------------------
SH_AOP = ('Phased Opp, Budget, Rev Plan', 7, 2, 56)   # sheet, header row, first col, n cols
SH_PWIN = ('Pursuits Pwins', 1, 1, 26)
SH_PWIN_DEP = ('Pursuits Pwins Dependent', 1, 1, 25)
SH_STAFF = ('StaffingData', 1, 1, 98)
# Dashboard Table5 -- annual operating plan. Read the TABLE, not the display
# grid on C3:G8: the table carries every year (2026-2033 here, unsorted),
# while the grid shows only five from the planning year. The schema has no
# five-year ceiling, so take them all.
SH_PLAN = ('Dashboard', 1, 27, 8)
# Dashboard!B1 holds the planning year. year_offset 1 == this year, so
# calendar_year = planning_year + year_offset - 1. Without it every
# year-based rollup silently returns nothing.
CELL_PLANNING_YEAR = ('Dashboard', 1, 2)

# Questionnaire column -> question code
Q_MAP = OrderedDict([
    ('TM 1a', 'TM1A'), ('TM 1b', 'TM1B'), ('TM 2', 'TM2'), ('TM 3', 'TM3'),
    ('TM 4', 'TM4'), ('TM 5', 'TM5'), ('PP 1', 'PP1'), ('P 1', 'P1'), ('P 2', 'P2'),
])
Q_NUMERIC = {'Invest %': 'INVEST_PCT'}

# Staffing sheet short names -> labor_category.code
CAT_SHORT = ['CM', 'TechLead', 'BDGen', 'PropMgr', 'VolLeads', 'Writers',
             'SMEEng', 'SMEOps', 'SMEProd', 'MatlMgr', 'PriceLead', 'Pricing',
             'Graphics', 'Compliance', 'RevBlue', 'RevPink', 'RevRed', 'RevGold']
CAT_CODE = {s: s.upper() for s in CAT_SHORT}

PHASE_SHORT = OrderedDict([('Strat', 'STRAT'), ('Sol', 'SOL'),
                           ('PreProp', 'PREPROP'), ('Final', 'FINAL'), ('EN', 'EN')])
WKS_COL = {'Wks_Strategy': 'STRAT', 'Wks_Solutioning': 'SOL',
           'Wks_PreProp': 'PREPROP', 'Wks_Final': 'FINAL', 'Wks_EN': 'EN'}

BID_MAP = {'Bid': 'BID', 'No Bid': 'NO_BID'}
OUTCOME_MAP = {'Won': 'WON', 'Lost': 'LOST', 'Canceled': 'CANCELLED',
               'Cancelled': 'CANCELLED'}
STAGE_MAP = {'Pre-BH': 'PRE_BH', 'Post-BH': 'POST_BH', 'Post-PTW': 'POST_PTW'}
CONTRACT_MAP = {'Cost Plus': 'COST_PLUS', 'Time & Materials': 'T_AND_M',
                'Fixed Price': 'FIXED_PRICE'}
OPPTYPE_MAP = {
    'Existing product solution, little modification required': 'EXIST_PROD',
    'Developmental/New Product': 'DEV_NEW',
    'Engineering/Technical Services': 'ENG_TECH_SVC',
    'Sustainment/O&M': 'SUSTAIN_OM',
}

MAX_YEARS = 5   # workbook ceiling; DB has none


# ---------------------------------------------------------------------
def read_table(wb, spec, limit=500, stop_on_blank=True):
    """Read a header row plus its data rows.

    stop_on_blank=False scans the whole range instead of stopping at the
    first empty row. Dashboard Table5 needs this: its header is on row 1
    but the data does not start until row 22, so stopping at the first
    blank silently returns nothing.
    """
    sheet, hdr_row, first_col, ncols = spec
    ws = wb[sheet]
    headers = [ws.cell(hdr_row, c).value for c in range(first_col, first_col + ncols)]
    out = []
    r = hdr_row + 1
    end = min(hdr_row + limit, ws.max_row + 1) if not stop_on_blank else hdr_row + limit
    while r < end:
        vals = [ws.cell(r, c).value for c in range(first_col, first_col + ncols)]
        if all(v is None for v in vals):
            if stop_on_blank:
                break
            r += 1
            continue
        out.append(dict(zip(headers, vals)))
        r += 1
    return out


def find_col(headers, prefix, year):
    """AOP year columns are named e.g. 'Probabilistic Revenue\nYear 1'."""
    want = f'{prefix}\nYear {year}'
    return want if want in headers else None


class Issue:
    def __init__(self):
        self.items = []

    def add(self, severity, where, msg):
        self.items.append((severity, where, msg))

    def report(self):
        if not self.items:
            print('No data issues found.')
            return 0
        errs = sum(1 for s, _, _ in self.items if s == 'ERROR')
        print(f'\n--- DATA ISSUES ({len(self.items)}: {errs} ERROR) ---')
        for sev, where, msg in self.items:
            print(f'  [{sev}] {where}: {msg}')
        return errs


# ---------------------------------------------------------------------
def extract(path, issues):
    wb = openpyxl.load_workbook(path, data_only=True)

    aop_rows = read_table(wb, SH_AOP)
    pwin_rows = read_table(wb, SH_PWIN)
    dep_rows = read_table(wb, SH_PWIN_DEP)
    staff_rows = read_table(wb, SH_STAFF)
    plan_rows = read_table(wb, SH_PLAN, stop_on_blank=False)

    sheet, row, col = CELL_PLANNING_YEAR
    planning_year = wb[sheet].cell(row, col).value
    if not isinstance(planning_year, int):
        issues.add('ERROR', 'planning year',
                   f'Dashboard!B1 is {planning_year!r}, expected a year. '
                   f'Year projections cannot be dated without it.')
        planning_year = None

    def keyed(rows, label):
        d = {}
        for r in rows:
            u = r.get('UID')
            if isinstance(u, int):
                if u in d:
                    issues.add('ERROR', f'{label} UID {u}', 'duplicate UID')
                d[u] = r
            elif u is not None:
                issues.add('WARN', label,
                           f'row with non-integer UID {u!r} skipped '
                           f'(name={r.get("Opportunity Name")!r})')
            else:
                name = r.get('Opportunity Name')
                if name:
                    issues.add('ERROR', label,
                               f'row with NO UID but name {name!r} -- orphan, skipped')
        return d

    aop = keyed(aop_rows, 'AOP')
    pwin = keyed(pwin_rows, 'Pwin')
    dep = keyed(dep_rows, 'Pwin_Dep')
    staff = keyed(staff_rows, 'StaffingData')

    # Cross-sheet integrity
    for u in sorted(set(pwin) - set(aop)):
        issues.add('ERROR', f'UID {u}', 'in Pwin but not AOP')
    for u in sorted(set(aop) - set(pwin)):
        issues.add('WARN', f'UID {u}', 'in AOP but not Pwin -- no assessment')
    for u in sorted(set(staff) - set(aop)):
        issues.add('ERROR', f'UID {u}', 'in StaffingData but not AOP')
    for u in sorted(set(dep) - set(pwin)):
        issues.add('ERROR', f'UID {u}', 'in Pwin_Dep but not Pwin')

    # Dependency targets must resolve
    for u, r in pwin.items():
        dep_uid = r.get('Dependent UID')
        if dep_uid is not None:
            if dep_uid not in aop:
                issues.add('ERROR', f'UID {u}',
                           f'Dependent UID {dep_uid} does not exist')
            if u not in dep:
                issues.add('WARN', f'UID {u}',
                           'has Dependent UID but no Pwin_Dep assessment row')
        elif u in dep:
            issues.add('WARN', f'UID {u}',
                       'has Pwin_Dep row but no Dependent UID on Pwin sheet')

    # Reference-value validation
    for u, r in aop.items():
        ot = r.get('Opportunity Type')
        if ot and ot not in OPPTYPE_MAP:
            issues.add('ERROR', f'UID {u}', f'unknown Opportunity Type {ot!r}')
        ph = r.get('Phase')
        if ph and ph not in STAGE_MAP:
            issues.add('ERROR', f'UID {u}', f'unknown Phase {ph!r}')
        bd = r.get('Bid/NoBid')
        if bd and bd not in BID_MAP:
            issues.add('ERROR', f'UID {u}', f'unknown Bid/NoBid {bd!r}')
        wl = r.get('Won/Lost')
        if wl and wl not in ('--',) and wl not in OUTCOME_MAP:
            issues.add('ERROR', f'UID {u}', f'unknown Won/Lost {wl!r}')
    for u, r in pwin.items():
        ct = r.get('Contract Type')
        if ct and ct not in CONTRACT_MAP:
            issues.add('ERROR', f'UID {u}', f'unknown Contract Type {ct!r}')

    # MARKET AUTHORITY: the Pwin sheet, not AOP.
    # The two sheets disagree on most rows in the demo workbook, and the
    # Pwin sheet's value is the one the ENGINE scored against -- so it is
    # the only value consistent with the stored Pwin. Taking AOP's would
    # produce a market breakdown that does not reconcile with the Pwins.
    for u, r in pwin.items():
        if not r.get('Market'):
            issues.add('ERROR', f'UID {u}', 'no Market on the Pwin sheet')
    a_mk = {u: r.get('Market') for u, r in aop.items()}
    mism = [u for u in pwin if u in a_mk and a_mk[u] != pwin[u].get('Market')]
    if mism:
        issues.add('WARN', 'market',
                   f'{len(mism)} pursuit(s) disagree between AOP and Pwin sheets; '
                   f'using the Pwin sheet value')
    blank_aop = [u for u, m in a_mk.items() if not m]
    if blank_aop:
        issues.add('WARN', 'market',
                   f'{len(blank_aop)} pursuit(s) have no Market on AOP; '
                   f'resolved from the Pwin sheet')

    markets = sorted({r.get('Market') for r in pwin.values() if r.get('Market')})

    # --- annual operating plan -------------------------------------
    plan = {}
    for r in plan_rows:
        y = r.get('Year')
        if not isinstance(y, int):
            continue
        vals = {'esc': r.get('EscalationRate'),
                'rev': r.get('Revenue Target'),
                'fee': r.get('Fee Target'),
                'bp': r.get('Budgeted B&P'),
                'inv': r.get('Budgeted Investment'),
                'ccr': r.get('Current Contract Revenue'),
                'ccf': r.get('Current Contract Fee')}
        # A row of zeros is a placeholder, not a plan. Skip it and say so.
        if not any(vals[k] for k in ('rev', 'fee', 'bp', 'inv')):
            issues.add('WARN', f'plan {y}', 'all targets zero or blank -- skipped')
            continue
        if y in plan:
            issues.add('ERROR', f'plan {y}', 'duplicate year in Dashboard table')
        plan[y] = vals

    if not plan:
        issues.add('WARN', 'plan', 'no annual targets found -- dashboard '
                                   'variances will be meaningless')
    else:
        yrs = sorted(plan)
        gaps = [y for y in range(yrs[0], yrs[-1] + 1) if y not in plan]
        if gaps:
            issues.add('WARN', 'plan',
                       f'missing year(s) {gaps} between {yrs[0]} and {yrs[-1]}')

    return {'aop': aop, 'pwin': pwin, 'dep': dep, 'staff': staff, 'plan': plan,
            'planning_year': planning_year,
            'markets': markets, 'aop_headers': list(aop_rows[0]) if aop_rows else []}


# ---------------------------------------------------------------------
def market_code(name):
    """Deterministic code from a market display name. Codes are IMMUTABLE
    once created, so this must stay stable."""
    out = []
    for ch in name.upper():
        out.append(ch if ch.isalnum() else '_')
    code = ''.join(out)
    while '__' in code:
        code = code.replace('__', '_')
    return code.strip('_')


def load(conn, data, args, issues):
    import psycopg
    cur = conn.cursor()

    # --- client
    cur.execute("INSERT INTO client (code, name) VALUES (%s,%s) "
                "ON CONFLICT (code) DO UPDATE SET name=EXCLUDED.name RETURNING id",
                (args.client_code, args.client_name))
    client_id = cur.fetchone()[0]

    # --- org: business -> business_unit
    cur.execute("INSERT INTO org_node (client_id,parent_id,node_type,code,name) "
                "VALUES (%s,NULL,'business',%s,%s) "
                "ON CONFLICT (client_id,code) DO UPDATE SET name=EXCLUDED.name "
                "RETURNING id",
                (client_id, 'BUSINESS', args.business))
    biz_id = cur.fetchone()[0]

    cur.execute("INSERT INTO org_node (client_id,parent_id,node_type,code,name,"
                "is_license_boundary) VALUES (%s,%s,'business_unit',%s,%s,TRUE) "
                "ON CONFLICT (client_id,code) DO UPDATE SET name=EXCLUDED.name "
                "RETURNING id",
                (client_id, biz_id, 'BU', args.business_unit))
    bu_id = cur.fetchone()[0]

    # --- markets (CDA-provisioned in production; derived here for migration)
    market_ids = {}
    for i, name in enumerate(data['markets'], 1):
        code = market_code(name)
        cur.execute("INSERT INTO market (client_id,code,name,display_order) "
                    "VALUES (%s,%s,%s,%s) ON CONFLICT (client_id,code) DO NOTHING "
                    "RETURNING id", (client_id, code, name, i))
        row = cur.fetchone()
        if row is None:
            cur.execute("SELECT id FROM market WHERE client_id=%s AND code=%s",
                        (client_id, code))
            row = cur.fetchone()
        market_ids[name] = row[0]

    # --- lookup caches
    def lookup(table, col='code'):
        cur.execute(f"SELECT {col}, id FROM {table}")
        return dict(cur.fetchall())

    opp_ids = lookup('opportunity_type')
    ctr_ids = lookup('contract_type')
    stage_ids = lookup('pipeline_stage')
    phase_ids = lookup('phase')
    cat_ids = lookup('labor_category')

    cur.execute("SELECT id FROM questionnaire_version WHERE code='pwin' AND is_active")
    qv_id = cur.fetchone()[0]

    cur.execute("SELECT q.code, q.id FROM question q WHERE q.questionnaire_version_id=%s",
                (qv_id,))
    q_ids = dict(cur.fetchall())

    # option lookup: (question_code, engine_value) -> option_id
    cur.execute("""SELECT q.code, o.engine_value, o.id
                     FROM question_option o
                     JOIN question q ON q.id=o.question_id
                    WHERE q.questionnaire_version_id=%s""", (qv_id,))
    opt_ids = {(qc, ev): oid for qc, ev, oid in cur.fetchall()}

    aop, pwin, dep, staff = data['aop'], data['pwin'], data['dep'], data['staff']
    headers = data['aop_headers']

    # --- pass 1: pursuits (no dependency yet -- targets may not exist)
    pursuit_ids = {}
    for uid in sorted(aop):
        a = aop[uid]
        p = pwin.get(uid, {})
        wl = a.get('Won/Lost')
        outcome = OUTCOME_MAP.get(wl) if wl and wl != '--' else None
        if outcome is None and a.get('Cancel Date'):
            outcome = 'CANCELLED'

        cur.execute("""
            INSERT INTO pursuit (
                client_id, org_node_id, external_opportunity_id, name,
                market_id, opportunity_type_id, contract_type_id, pipeline_stage_id,
                is_sole_source, bidders, bid_decision,
                planned_total_award_value, planned_fee_rate, investment_pct,
                min_bp, max_bp, planned_investment,
                bp_start_date, proposal_due_date, contract_award_date,
                period_end_date, stage_completed_date, cancel_date,
                outcome, crm_pwin, black_hat_ptw_complete)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                    %s,%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id""",
            (client_id, bu_id,
             str(a.get('Opportunity ID')) if a.get('Opportunity ID') is not None else None,
             a.get('Opportunity Name') or f'UID {uid}',
             market_ids.get(p.get('Market') or a.get('Market')),
             opp_ids.get(OPPTYPE_MAP.get(a.get('Opportunity Type'))),
             ctr_ids.get(CONTRACT_MAP.get(p.get('Contract Type'))),
             stage_ids.get(STAGE_MAP.get(a.get('Phase'))),
             bool(a.get('Sole Source')),
             p.get('Bidders'),
             'BID' if args.force_bid else BID_MAP.get(a.get('Bid/NoBid')),
             a.get('Planned Total Award Value'), a.get('Planned Fee'),
             p.get('Invest %'),
             a.get('Min B&P'), a.get('Max B&P'), a.get('Planned Investment'),
             a.get('B&P Start'), a.get('Proposal Due'), a.get('Contract Award'),
             a.get('End Date'), a.get('Date Phase Completed'), a.get('Cancel Date'),
             outcome,
             parse_pct(a.get('SalesForce Pwin')),
             bool(p.get('BHPTWComplete'))))
        pursuit_ids[uid] = cur.fetchone()[0]

    # --- pass 2: dependencies
    for uid, p in pwin.items():
        dep_uid = p.get('Dependent UID')
        if dep_uid and dep_uid in pursuit_ids and uid in pursuit_ids:
            cur.execute("UPDATE pursuit SET depends_on_pursuit_id=%s WHERE id=%s",
                        (pursuit_ids[dep_uid], pursuit_ids[uid]))

    # --- assessments + answers
    def write_assessment(uid, row, scenario, blended=None):
        pid = pursuit_ids[uid]
        cur.execute("""
            INSERT INTO pwin_assessment (
                pursuit_id, questionnaire_version_id, scenario, engine_version,
                pwin, base_pwin, blended_pwin,
                score_tech, score_mgmt, score_past_perf,
                price_position, competitor_price_position, is_current)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,TRUE)
            RETURNING id""",
            (pid, qv_id, scenario, args.engine_version,
             row.get('Pwin'), row.get('Base Pwin'), blended,
             row.get('Tech'), row.get('Mgmt'), row.get('PP'),
             row.get('Price'), row.get('C Price')))
        aid = cur.fetchone()[0]

        for col, qcode in Q_MAP.items():
            val = row.get(col)
            if val is None:
                continue
            oid = opt_ids.get((qcode, val))
            if oid is None:
                issues.add('ERROR', f'UID {uid} {scenario}',
                           f'{qcode}: no option matching {val!r}')
                continue
            cur.execute("INSERT INTO pwin_answer (pwin_assessment_id,question_id,"
                        "question_option_id) VALUES (%s,%s,%s)",
                        (aid, q_ids[qcode], oid))

        for col, qcode in Q_NUMERIC.items():
            val = row.get(col)
            if val is not None:
                cur.execute("INSERT INTO pwin_answer (pwin_assessment_id,question_id,"
                            "numeric_value) VALUES (%s,%s,%s)",
                            (aid, q_ids[qcode], val))
        return aid

    for uid in sorted(pwin):
        if uid not in pursuit_ids:
            continue
        base = pwin[uid]
        # When a DEPENDENT_WON assessment exists, the Pwin on the main sheet
        # is the BLENDED result and Base Pwin is the standalone value.
        blended = base.get('Pwin') if uid in dep else None
        write_assessment(uid, base, 'BASE', blended)
        if uid in dep:
            write_assessment(uid, dep[uid], 'DEPENDENT_WON')

    # --- year projections
    for uid, a in aop.items():
        pid = pursuit_ids[uid]
        for y in range(1, MAX_YEARS + 1):
            vals = {
                'billable_contract_days': a.get(find_col(headers, 'Billable Contract Days', y)),
                'probabilistic_revenue': a.get(find_col(headers, 'Probabilistic Revenue', y)),
                'probabilistic_fee': a.get(find_col(headers, 'Probabilistic Fee', y)),
                'bp_days': a.get(find_col(headers, 'B&P Days', y)),
                'bp_required': a.get(find_col(headers, 'B&P Required', y)),
                'planned_investment': a.get(f'Planned Investment Year {y}'),
            }
            if all(v in (None, 0) for v in vals.values()):
                continue
            cal = (data['planning_year'] + y - 1) if data['planning_year'] else None
            cur.execute("""
                INSERT INTO pursuit_year_projection
                    (pursuit_id, year_offset, calendar_year,
                     billable_contract_days,
                     probabilistic_revenue, probabilistic_fee, bp_days,
                     bp_required, planned_investment)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                (pid, y, cal, vals['billable_contract_days'],
                 vals['probabilistic_revenue'], vals['probabilistic_fee'],
                 vals['bp_days'], vals['bp_required'], vals['planned_investment']))

    # --- staffing
    for uid, s in staff.items():
        if uid not in pursuit_ids:
            continue
        pid = pursuit_ids[uid]
        for wcol, pcode in WKS_COL.items():
            w = s.get(wcol)
            if w is not None:
                cur.execute("INSERT INTO pursuit_phase_duration "
                            "(pursuit_id,phase_id,weeks) VALUES (%s,%s,%s)",
                            (pid, phase_ids[pcode], w))
        for cshort in CAT_SHORT:
            for pshort, pcode in PHASE_SHORT.items():
                col = f'FTE_{cshort}_{pshort}'
                v = s.get(col)
                if v is None:
                    continue
                cur.execute("INSERT INTO pursuit_staffing "
                            "(pursuit_id,labor_category_id,phase_id,fte) "
                            "VALUES (%s,%s,%s,%s)",
                            (pid, cat_ids[CAT_CODE[cshort]], phase_ids[pcode], v))
        cur.execute("""INSERT INTO pursuit_staffing_meta
                       (pursuit_id, effective_bp_pct, engine_version)
                       VALUES (%s,%s,%s)""",
                    (pid, s.get('Effective_BP_Pct'), args.engine_version))

    # --- annual operating plan (targets the pipeline is measured against)
    for y, v in sorted(data['plan'].items()):
        cur.execute("""
            INSERT INTO plan_year (org_node_id, calendar_year, escalation_rate,
                revenue_target, fee_target, budgeted_bp, budgeted_investment,
                current_contract_revenue, current_contract_fee)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT (org_node_id, calendar_year) DO UPDATE SET
                escalation_rate          = EXCLUDED.escalation_rate,
                revenue_target           = EXCLUDED.revenue_target,
                fee_target               = EXCLUDED.fee_target,
                budgeted_bp              = EXCLUDED.budgeted_bp,
                budgeted_investment      = EXCLUDED.budgeted_investment,
                current_contract_revenue = EXCLUDED.current_contract_revenue,
                current_contract_fee     = EXCLUDED.current_contract_fee""",
            (bu_id, y, v['esc'], v['rev'], v['fee'], v['bp'], v['inv'],
             v['ccr'], v['ccf']))

    conn.commit()
    return pursuit_ids


def parse_pct(v):
    """SalesForce Pwin arrives as '30%' / '60%' / '90%' or None."""
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip().rstrip('%')
        if not v:
            return None
        try:
            return float(v) / 100.0
        except ValueError:
            return None
    return float(v)


# ---------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--workbook', required=True)
    ap.add_argument('--dsn')
    ap.add_argument('--dry-run', action='store_true')
    ap.add_argument('--client-code', default='DEMO')
    ap.add_argument('--client-name', default='Demo Client')
    ap.add_argument('--business', default='Demo Business')
    ap.add_argument('--business-unit', default='Demo Business Unit')
    ap.add_argument('--engine-version', default='0.23')
    ap.add_argument('--force-bid', action='store_true',
                    help='Load every pursuit as Bid, ignoring the workbook '
                         'Bid/NoBid column. Narrowing the pipeline is the '
                         "tool's job in the demo, not the data's.")
    args = ap.parse_args()

    issues = Issue()
    data = extract(args.workbook, issues)

    print(f'Parsed: {len(data["aop"])} AOP, {len(data["pwin"])} Pwin, '
          f'{len(data["dep"])} Pwin_Dep, {len(data["staff"])} StaffingData')
    deps = [(u, r.get("Dependent UID")) for u, r in data['pwin'].items()
            if r.get('Dependent UID')]
    print(f'Dependencies: {deps}')
    py = data.get('planning_year')
    print(f'Planning year: {py}  -> projections dated '
          f'{py}-{py+4}' if py else 'Planning year: MISSING')
    nb = sum(1 for r in data['aop'].values() if r.get('Bid/NoBid') == 'No Bid')
    if args.force_bid and nb:
        print(f'--force-bid: {nb} No Bid pursuit(s) will load as Bid')
    mk = sorted({r.get('Market') for r in data['pwin'].values() if r.get('Market')})
    print(f'Markets (from Pwin sheet): {mk}')
    pl = data['plan']
    if pl:
        yrs = sorted(pl)
        print(f'Plan years: {len(pl)} ({yrs[0]}-{yrs[-1]})  '
              f'revenue targets {min(pl[y]["rev"] for y in yrs):,.0f}'
              f'-{max(pl[y]["rev"] for y in yrs):,.0f}')
        ccr = sum(1 for y in yrs if pl[y]['ccr'])
        print(f'  other contract revenue populated in {ccr} of {len(pl)} years')

    errs = issues.report()

    if args.dry_run:
        print('\n--dry-run: database not touched.')
        return 0 if errs == 0 else 1

    if not args.dsn:
        print('ERROR: --dsn required unless --dry-run', file=sys.stderr)
        return 2
    if errs:
        print(f'\nRefusing to load: {errs} ERROR-level issue(s). '
              f'Fix the workbook or override deliberately.', file=sys.stderr)
        return 1

    import psycopg
    with psycopg.connect(args.dsn) as conn:
        pids = load(conn, data, args, issues)
        print(f'\nLoaded {len(pids)} pursuits.')
        issues.report()
    return 0


if __name__ == '__main__':
    sys.exit(main())
