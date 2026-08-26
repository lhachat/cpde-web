#!/usr/bin/env python3
"""
gen_demo.py -- generate a synthetic CPDE demo pipeline.

Design goals:
  * LEVEL LOADED  -- B&P demand roughly flat across the 5-year window rather
                     than front- or back-weighted.
  * 3x OVERSUBSCRIBED -- total B&P demand each year is ~3x the B&P budget,
                     so the portfolio cannot be pursued as-is. This is the
                     whole point: the tool exists to choose.
  * NO REAL DATA  -- markets, names, values and dates are invented.

Outputs:
  demo_pipeline.xlsx  -- import-ready for the CPDE workbook (input columns
                         only; the workbook recalculates everything derived)
  demo_seed.sql       -- loads the same pipeline into the v2 Postgres schema

Usage:
  python gen_demo.py --pursuits 96 --seed 7
"""
import argparse, random, datetime, math, json

PY = 2026                     # planning year
YEARS = [PY + i for i in range(5)]
TODAY = datetime.date(2026, 8, 19)
TODAY = datetime.date(2026, 8, 19)

# --- plan targets (the constraint the pipeline is measured against) ---
# Company shooting for $250M/yr AOP revenue. B&P budget is 0.9% of revenue
# target (industry-typical). Escalation 2.8%.
_REV = {2026: 250_000_000, 2027: 258_000_000, 2028: 266_000_000,
        2029: 274_000_000, 2030: 282_000_000}
# B&P budget = 2% of revenue target. That 2% already carries the win-rate
# assumption: ~0.9% to bid a pursuit, and you win roughly half of them.
# Investment budget = 0.5% -- investment is rare and deliberate.
TARGETS = {y: dict(esc=0.028, rev=_REV[y], fee=round(_REV[y] * 0.075),
                   bp=round(_REV[y] * 0.020), inv=round(_REV[y] * 0.005))
           for y in _REV}
TARGET_AOP = 250_000_000      # annual AOP revenue the company is shooting for
ASSUMED_PWIN = 0.30           # proxy only -- the engine computes the real value

# Revenue under contract TODAY, rolling off as awards expire. A company at
# $250M/yr does not get it from the pipeline -- the pipeline fills the gap
# that opens as the existing base declines. This is what makes the near-year
# shortfall real: 2026 is nearly covered, 2030 is not.
# Legacy contracts predating this pipeline. DERIVED, not assumed: it is
# whatever is left after won pursuits and the Pwin-weighted open pipeline,
# so total revenue tracks COVERAGE below. A capture organisation is
# comfortable in the near years and exposed in the out years -- that gap is
# the problem the tool exists to solve.
# Legacy contracts predating the pipeline, rolling off linearly to zero.
CURRENT_BASE = {2026: 91_000_000, 2027: 68_000_000, 2028: 45_000_000,
                2029: 22_000_000, 2030: 0}

# The demo's whole point: the pipeline is roughly TWICE what is needed, on
# BOTH constraints at once. Culling about half lands the portfolio on plan.
# Needed revenue = target - legacy base. Pipeline should deliver 2x that.
# The pipeline holds ~3x what can actually be pursued. That tension is the
# point: the tool exists to narrow it.
PIPELINE_MULTIPLE = 3.0
BP_MULTIPLE = 3.0
WIN_RATE = 0.50          # applies to DECIDED pursuits (deliberately selected)
TARGET_OPEN_PWIN = 0.30  # raw pipeline average, before any prioritisation
CURRENT_FEE_RATE = 0.079
OVERSUBSCRIPTION = BP_MULTIPLE

MARKETS = [
    ("GSS",   "Ground Systems & Sensors"),
    ("ACP",   "Airborne Comms & Processing"),
    ("MSN",   "Mission Networks"),
    ("SUS",   "Sustainment & Depot"),
    ("SPC",   "Space Payloads"),
    ("TRN",   "Training & Simulation"),
]
OPP_TYPES = [
    ("Existing product solution, little modification required", 0.28),
    ("Developmental/New Product",                               0.22),
    ("Engineering/Technical Services",                          0.32),
    ("Sustainment/O&M",                                         0.18),
]
CONTRACTS = [("Cost Plus", 0.34), ("Fixed Price", 0.42), ("Time & Materials", 0.24)]
STAGES = [("Pre-BH", 0.62), ("Post-BH", 0.26), ("Post-PTW", 0.12)]

ANSWERS = {
    "TM 1a": {"PRODUCT": ["Same", "Better", "Worse"],
              "SERVICES": ["On contract today",
                           "Yes, but not as many building blocks as competitor",
                           "No, but our teammates are on contract today", "No"]},
    "TM 1b": {"PRODUCT": ["Same Level", "More Mature", "Less Mature"],
              "SERVICES": ["On contract today",
                           "Yes, but not as many building blocks as us", "No"]},
    "TM 2": ["Yes, us", "Yes, one of the competitors", "No"],
    "TM 3": ["We are performing satisfactorily/unknown", "We have performance issues",
             "N/A", "Incumbent competitor is performing satisfactorily/unknown",
             "Incumbent competitor has performance issues"],
    "TM 4": ['Yes, we will outsource most of the actual work requested ("noble work")',
             'Yes, we will outsource some of the actual work requested ("noble work")', "No"],
    "TM 5": ["No", "Low", "Moderate", "High"],
    "PP 1": ["Yes", "No"],
    "P 1": ["3% above normal", "2% above normal", "1% above normal", "Normal Bid",
            "1% lower than normal", "2% lower than normal", "3% lower than normal",
            "4% lower than normal"],
    "P 2": ["Best Value", "LPTA"],
}
# Weighted so the pipeline isn't uniformly mediocre — a realistic spread.
# Weighted so the RAW pipeline averages ~30% Pwin -- it is full of pursuits
# that should not be bid. Selection is the tool's job, not the data's.
W_TM2 = [0.12, 0.46, 0.42]          # usually a competitor incumbent, or none
W_TM3 = [0.16, 0.12, 0.22, 0.38, 0.12]
W_TM4 = [0.18, 0.34, 0.48]
W_TM5 = [1.0, 0.0, 0.0, 0.0]   # no investment by default; assigned deliberately
W_PP1 = [0.26, 0.74]
# Off-normal bidding is RARE -- ~3% of pursuits. Bidding aggressively is a
# lever the demo pulls deliberately; it should not be background variation.
W_P1 = [0.002, 0.004, 0.008, 0.970, 0.008, 0.005, 0.002, 0.001]
W_P1_HIGH = [0.004, 0.008, 0.016, 0.972]  # GSS: Normal Bid or above only
W_P2 = [0.78, 0.22]

NOUNS = ["Sentinel", "Vanguard", "Meridian", "Aperture", "Cascade", "Harbor", "Lattice",
         "Beacon", "Quarry", "Trestle", "Foundry", "Anvil", "Cirrus", "Basalt", "Kestrel",
         "Pinnacle", "Trellis", "Ironwood", "Halyard", "Verdigris", "Cobalt", "Marlin",
         "Ridgeline", "Sable", "Tessera", "Windlass", "Zephyr", "Alder", "Bulwark",
         "Cordite", "Drumlin", "Ember", "Fathom", "Granite", "Hollow", "Inlet"]
QUALS = ["Program", "Recompete", "Follow-On", "Phase II", "Modernization", "Refresh",
         "Bridge", "Expansion", "IDIQ", "Task Order", "Upgrade", "Sustainment"]


def pick(pairs, rng):
    r, c = rng.random(), 0.0
    for v, w in pairs:
        c += w
        if r <= c:
            return v
    return pairs[-1][0]


def wpick(opts, weights, rng):
    r, c = rng.random(), 0.0
    for v, w in zip(opts, weights):
        c += w
        if r <= c:
            return v
    return opts[-1]


def tm3_for(tm2, rng):
    """TM3 options are fully determined by TM2 (PwinForm.CB_TM2_Change)."""
    if tm2 == "Yes, us":
        return wpick(["We are performing satisfactorily/unknown",
                      "We have performance issues"], [0.80, 0.20], rng)
    if tm2 == "Yes, one of the competitors":
        return wpick(["Incumbent competitor is performing satisfactorily/unknown",
                      "Incumbent competitor has performance issues"], [0.72, 0.28], rng)
    return "N/A"


# TM3 options are fully determined by TM2 (PwinForm CB_TM2_Change).
TM3_BY_TM2 = {
    "Yes, us": (["We are performing satisfactorily/unknown",
                 "We have performance issues"], [0.72, 0.28]),
    "Yes, one of the competitors": (
        ["Incumbent competitor is performing satisfactorily/unknown",
         "Incumbent competitor has performance issues"], [0.68, 0.32]),
    "No": (["N/A"], [1.0]),
}


def tm3_for(tm2, rng):
    opts, w = TM3_BY_TM2[tm2]
    return wpick(opts, w, rng)


def enforce_gates(r):
    """Re-apply every PwinForm gate after any answer is changed."""
    prod = is_product(r["otype"])
    a = r["ans"]
    if not prod:
        if a["TM 1a"] == "No" and a["TM 1b"] == "Yes, but not as many building blocks as us":
            a["TM 1b"] = "No"
        if a["TM 1a"] == "No" and a["TM 1b"] == "No":
            a["TM 2"] = "No"
    if a["TM 3"] not in TM3_BY_TM2[a["TM 2"]][0]:
        a["TM 3"] = TM3_BY_TM2[a["TM 2"]][0][0]
    return r


def is_product(t):
    t = t.lower()
    return "product" in t or "dev" in t or "existing" in t


def build(n, seed):
    rng = random.Random(seed)
    used = set()
    rows = []

    # Level-load: distribute B&P start dates uniformly across the window,
    # starting one year before the planning year so 2026 has in-flight work.
    win_start = datetime.date(PY - 2, 1, 1)
    win_end = datetime.date(PY + 4, 11, 30)
    span = (win_end - win_start).days

    for i in range(n):
        # --- name
        while True:
            nm = f"{rng.choice(NOUNS)} {rng.choice(QUALS)}"
            if nm not in used:
                used.add(nm)
                break

        mkt = rng.choice(MARKETS)[0]
        otype = pick(OPP_TYPES, rng)
        ctype = pick(CONTRACTS, rng)
        prod = is_product(otype)

        # --- award value: log-normal, $1.5M .. ~$300M, median ~$18M
        v = math.exp(rng.gauss(math.log(42_000_000), 1.10))
        value = max(4_000_000, min(650_000_000, v))
        value = round(value / 100_000) * 100_000

        # --- fee by contract type
        fee = {"Cost Plus": rng.uniform(.055, .085),
               "Fixed Price": rng.uniform(.085, .145),
               "Time & Materials": rng.uniform(.055, .095)}[ctype]
        fee = round(fee, 4)

        # --- dates, spread evenly (this is what level-loads B&P)
        bp_start = win_start + datetime.timedelta(days=int(span * i / max(1, n - 1))
                                                  + rng.randint(-40, 40))
        capture_wks = rng.choice([14, 18, 22, 26, 30, 36])
        due = bp_start + datetime.timedelta(weeks=capture_wks)
        award = due + datetime.timedelta(days=rng.choice([90, 120, 150, 180, 240]))
        pop_years = rng.choice([1, 2, 3, 3, 4, 5, 5])
        end = award + datetime.timedelta(days=int(365.25 * pop_years))

        # --- B&P: scaled to value but bounded; ~0.9-2.4% of award
        bp = value * rng.uniform(.004, .009)
        bp = max(35_000, min(1_400_000, bp))

        # Sole source and "1 bidder" are the same fact. ~95% win rate: the
        # only way to lose is the customer not awarding at all.
        sole = rng.random() < 0.07
        # 3 bidders is the norm; 2 and 4 less common. 1 bidder ONLY when
        # sole source -- they are the same condition.
        bidders = 1 if sole else wpick([3, 4, 2], [0.74, 0.13, 0.13], rng)
        # A pursuit whose B&P has not started cannot have completed a Black
        # Hat or PTW. Anything future-dated is necessarily Pre-BH.
        if bp_start > TODAY:
            stage = "Pre-BH"
        else:
            elapsed = (TODAY - bp_start).days / max(1, (due - bp_start).days)
            if elapsed > 0.75:
                stage = wpick(["Post-PTW", "Post-BH", "Pre-BH"], [0.55, 0.33, 0.12], rng)
            elif elapsed > 0.40:
                stage = wpick(["Post-BH", "Pre-BH", "Post-PTW"], [0.52, 0.36, 0.12], rng)
            else:
                stage = wpick(["Pre-BH", "Post-BH"], [0.82, 0.18], rng)

        # --- outcome: only for pursuits whose award date has passed
        outcome, out_date = None, None
        decided = award < TODAY and bp_start < TODAY
        if decided:
            outcome = "PENDING"          # Won/Lost assigned after the fact
            out_date = award
            stage = "Post-PTW" if rng.random() < .78 else "Post-BH"


        # --- questionnaire (GATED -- mirrors PwinForm exactly) -----------
        key = "PRODUCT" if prod else "SERVICES"
        ans = {}

        # TM1a
        ans["TM 1a"] = (wpick(ANSWERS["TM 1a"]["PRODUCT"], [0.46, 0.16, 0.38], rng)
                        if prod else
                        wpick(ANSWERS["TM 1a"]["SERVICES"], [0.10, 0.26, 0.18, 0.46], rng))

        # TM1b -- UpdateTM1bOptions_: services, TM1a="No" removes
        # "Yes, but not as many building blocks as us" (we cannot have more
        # experience than competitors if we have never done the work).
        if prod:
            ans["TM 1b"] = wpick(ANSWERS["TM 1b"]["PRODUCT"], [0.40, 0.14, 0.46], rng)
        elif ans["TM 1a"] == "No":
            ans["TM 1b"] = wpick(["On contract today", "No"], [0.42, 0.58], rng)
        else:
            ans["TM 1b"] = wpick(ANSWERS["TM 1b"]["SERVICES"], [0.30, 0.32, 0.38], rng)

        # TM2 -- UpdateTM2Options_: services, if neither TM1a nor TM1b has a
        # "yes" then there is no incumbent to speak of; TM2 collapses to "No".
        if (not prod) and ans["TM 1a"] == "No" and ans["TM 1b"] == "No":
            ans["TM 2"] = "No"
        else:
            ans["TM 2"] = wpick(ANSWERS["TM 2"], W_TM2, rng)

        # TM3 -- CB_TM2_Change drives this list entirely.
        ans["TM 3"] = tm3_for(ans["TM 2"], rng)

        ans["TM 4"] = wpick(ANSWERS["TM 4"], W_TM4, rng)
        ans["TM 5"] = "No"          # investment assigned deliberately later
        ans["PP 1"] = wpick(ANSWERS["PP 1"], W_PP1, rng)
        ans["P 1"] = (wpick(ANSWERS["P 1"][:4], W_P1_HIGH, rng) if mkt == "GSS"
                      else wpick(ANSWERS["P 1"], W_P1, rng))
        ans["P 2"] = wpick(ANSWERS["P 2"], W_P2, rng)
        invest = 0.0

        rows.append(dict(
            idx=i + 1, opp_id=1000 + i, name=nm, market=mkt, otype=otype, ctype=ctype,
            value=value, fee=fee, bp=round(bp, 2), invest_pct=invest,
            invest=round(value * invest, 2),
            bp_start=bp_start, due=due, award=award, end=end,
            sole=sole, bidders=bidders, stage=stage, bid="Bid",
            outcome=outcome, out_date=out_date, ans=ans, dep=None))

    # --- dependencies: ~9% of pursuits depend on an earlier-awarding one
    pool = [r for r in rows if not r["outcome"]]
    pool.sort(key=lambda r: r["award"])
    ndep = max(4, int(len(rows) * 0.09))
    for r in rng.sample(pool[len(pool) // 3:], min(ndep, len(pool) // 3)):
        cands = [c for c in pool if c["award"] < r["bp_start"] and c is not r]
        if cands:
            r["dep"] = rng.choice(cands)["idx"]

    # --- targeted investment: rare, deliberate, and only where it makes
    # sense -- developmental work, or displacing a well-performing incumbent.
    # Mid-size pursuits only: investing 5% of a $400M pursuit is not a real
    # decision, it is the whole budget. Investment is a targeted bet.
    cands = [r for r in rows if not r["outcome"]
             and 8_000_000 <= r["value"] <= 70_000_000
             and (r["otype"] == "Developmental/New Product"
                  or r["ans"]["TM 3"] == "Incumbent competitor is performing satisfactorily/unknown")]
    cands.sort(key=lambda r: -r["value"])
    for r in cands[:rng.randint(9, 12)]:
        r["ans"]["TM 5"] = wpick(["Low", "Moderate", "High"], [0.55, 0.35, 0.10], rng)
        r["invest_pct"] = {"Low": 0.01, "Moderate": 0.02, "High": 0.05}[r["ans"]["TM 5"]]
        r["invest"] = round(r["value"] * r["invest_pct"], 2)

    # Every open pursuit is Bid. Narrowing the pipeline is the demo's job,
    # not something baked into the data.

    # --- assign Won/Lost to decided pursuits at WIN_RATE ---------------
    dec = [r for r in rows if r["outcome"] == "PENDING"]
    dec.sort(key=lambda r: r["award"])
    # Decided pursuits are older and smaller than what is in the pipeline now.
    for r in dec:
        r["value"] = round(r["value"] * 0.12 / 100_000) * 100_000
        r["bp"] = r["bp"] * 0.12
        r["invest"] = round(r["value"] * r["invest_pct"], 2)
    nwin = round(len(dec) * WIN_RATE)
    winners = set(rng.sample(range(len(dec)), nwin))
    for i, r in enumerate(dec):
        r["outcome"] = "Won" if i in winners else "Lost"
        # DECIDED pursuits were deliberately selected, so their profiles are
        # stronger than the raw pipeline average -- that is the whole point.
        # DECIDED pursuits were deliberately selected -> stronger profiles.
        # Gating rules still apply.
        prod = is_product(r["otype"])
        r["ans"]["PP 1"] = "No" if rng.random() < .88 else "Yes"
        r["ans"]["TM 4"] = wpick(ANSWERS["TM 4"], [0.06, 0.24, 0.70], rng)
        if prod:
            r["ans"]["TM 1a"] = wpick(ANSWERS["TM 1a"]["PRODUCT"], [0.34, 0.54, 0.12], rng)
            r["ans"]["TM 1b"] = wpick(ANSWERS["TM 1b"]["PRODUCT"], [0.34, 0.50, 0.16], rng)
            r["ans"]["TM 2"] = wpick(["Yes, us", "No", "Yes, one of the competitors"],
                                     [0.44, 0.34, 0.22], rng)
        else:
            a = wpick(ANSWERS["TM 1a"]["SERVICES"], [0.34, 0.30, 0.20, 0.16], rng)
            if a == "No":
                b = wpick(["On contract today", "No"], [0.70, 0.30], rng)
            else:
                b = wpick(ANSWERS["TM 1b"]["SERVICES"], [0.20, 0.30, 0.50], rng)
            r["ans"]["TM 1a"], r["ans"]["TM 1b"] = a, b
            r["ans"]["TM 2"] = ("No" if (a == "No" and b == "No")
                                else wpick(["Yes, us", "No", "Yes, one of the competitors"],
                                           [0.44, 0.34, 0.22], rng))
        r["ans"]["TM 3"] = tm3_for(r["ans"]["TM 2"], rng)
    return rows


def bp_by_year(rows):
    """Spread each pursuit's B&P across its capture window, by calendar year."""
    out = {y: 0.0 for y in YEARS}
    for r in rows:
        if r["bid"] != "Bid" or r["outcome"] in ("Canceled",):
            continue
        d0, d1 = r["bp_start"], r["due"]
        days = max(1, (d1 - d0).days)
        for y in YEARS:
            ys, ye = datetime.date(y, 1, 1), datetime.date(y, 12, 31)
            ov = (min(d1, ye) - max(d0, ys)).days
            if ov > 0:
                out[y] += r["bp"] * ov / days
    return out


_WON_REV = {y: 0.0 for y in YEARS}


def needed(y):
    """Revenue the OPEN pipeline must supply: target less legacy base less
    pursuits already won. Won work is secured revenue, not opportunity."""
    return max(0, TARGETS[y]["rev"] - CURRENT_BASE[y] - _WON_REV[y])


def rev_by_year(rows):
    """Estimate probabilistic revenue per calendar year.

    APPROXIMATION. Uses ASSUMED_PWIN rather than a real Pwin, because Pwin
    comes from the engine. Award value is spread evenly across the period of
    performance. Good enough to size the pipeline; verify after import."""
    out = {y: 0.0 for y in YEARS}
    for r in rows:
        if r["bid"] != "Bid":
            continue
        # LOST and CANCELED produce no revenue -- ever. WON is certain, so it
        # carries no probability weighting. Only open pursuits are Pwin-weighted.
        if r["outcome"] in ("Lost", "Canceled"):
            continue
        weight = 1.0 if r["outcome"] == "Won" else ASSUMED_PWIN
        d0, d1 = r["award"], r["end"]
        days = max(1, (d1 - d0).days)
        annual = r["value"] * weight
        for y in YEARS:
            ys, ye = datetime.date(y, 1, 1), datetime.date(y, 12, 31)
            ov = (min(d1, ye) - max(d0, ys)).days
            if ov > 0:
                out[y] += annual * ov / days
    return out


def scale_values(rows, target=None, iters=30):
    """Scale OPEN award values so the open pipeline runs at PIPELINE_MULTIPLE
    times what is still needed after legacy base and won work."""
    target = target or TARGET_AOP
    for y in YEARS:
        _WON_REV[y] = rev_by_year([r for r in rows if r["outcome"] == "Won"])[y]
    for _ in range(iters):
        cur = rev_by_year([r for r in rows if not r["outcome"]])
        # Anchor on the OUT-YEARS. That is where the existing base has rolled
        # off and the pipeline has to carry the plan by itself -- and it is the
        # gap a capture organisation is actually managing toward.
        anchor = YEARS
        mid = sum(cur[y] for y in anchor) / len(anchor)
        if mid <= 0:
            break
        gap = sum(needed(y) * PIPELINE_MULTIPLE for y in anchor) / len(anchor)
        f = (gap / mid) ** 0.5
        if abs(f - 1) < 0.01:
            break
        for r in rows:
            if r["outcome"]:          # decided pursuits are history; don't rescale
                continue
            r["value"] = round(max(4_000_000, min(420_000_000, r["value"] * f)) / 100_000) * 100_000
            r["invest"] = round(r["value"] * r["invest_pct"], 2)
    return rows


def rescale(rows, target_ratio=OVERSUBSCRIPTION, iters=14):
    """Scale B&P so each year's demand lands near target_ratio x budget."""
    for _ in range(iters):
        cur = bp_by_year(rows)
        worst = max(abs(cur[y] / (TARGETS[y]["bp"] * target_ratio) - 1) for y in YEARS)
        if worst < 0.04:
            break
        for r in rows:
            mid = r["bp_start"] + (r["due"] - r["bp_start"]) / 2
            y = min(YEARS, key=lambda yy: abs(yy - mid.year))
            want = TARGETS[y]["bp"] * target_ratio
            if cur[y] > 0:
                r["bp"] = max(35_000, min(1_600_000, r["bp"] * (want / cur[y]) ** 0.6))
    return rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--pursuits", type=int, default=96)
    ap.add_argument("--seed", type=int, default=7)
    ap.add_argument("--xlsx", default="demo_pipeline.xlsx")
    ap.add_argument("--sql", default="demo_seed.sql")
    ap.add_argument("--json", default="demo_payload.json")
    ap.add_argument("--target-aop", type=float, default=TARGET_AOP)
    ap.add_argument("--markets", default="demo_markets.json")
    args = ap.parse_args()

    rows = build(args.pursuits, args.seed)
    for y in YEARS:
        _WON_REV[y] = rev_by_year([r for r in rows if r["outcome"] == "Won"])[y]
    # B&P capacity is the binding constraint -- size the pipeline against it.
    rows = rescale(rows)

    bp = bp_by_year(rows)
    print(f"Generated {len(rows)} pursuits\n")
    print(f"{'Year':6}{'B&P budget':>14}{'B&P demand':>14}{'Ratio':>8}{'Concurrent':>12}")
    for y in YEARS:
        conc = sum(1 for r in rows
                   if r["bid"] == "Bid" and r["bp_start"].year <= y <= r["due"].year)
        print(f"{y:<6}{TARGETS[y]['bp']:>14,.0f}{bp[y]:>14,.0f}"
              f"{bp[y]/TARGETS[y]['bp']:>8.2f}x{conc:>11}")
    won_r = rev_by_year([r for r in rows if r["outcome"] == "Won"])
    open_r = rev_by_year([r for r in rows if not r["outcome"]])
    print(f"\n{'Year':6}{'Target':>13}{'Legacy':>12}{'Needed':>13}{'Won':>13}"
          f"{'Open@Pwin':>13}{'Total rev':>13}{'open/need':>10}")
    for y in YEARS:
        nd = needed(y)
        print(f"{y:<6}{TARGETS[y]['rev']:>13,.0f}{CURRENT_BASE[y]:>12,.0f}"
              f"{nd:>13,.0f}{won_r[y]:>13,.0f}{open_r[y]:>13,.0f}"
              f"{won_r[y]+open_r[y]+CURRENT_BASE[y]:>13,.0f}"
              f"{(open_r[y]/nd if nd else 0):>9.2f}x")
    print(f"  (pipeline estimate assumes Pwin={ASSUMED_PWIN:.0%}; the engine computes the real value)")
    print()
    won = sum(1 for r in rows if r["outcome"] == "Won")
    lost = sum(1 for r in rows if r["outcome"] == "Lost")
    print(f"\nOutcomes: {won} won, {lost} lost, {len(rows)-won-lost} open"
          f"   win rate {won/(won+lost):.1%}")
    gss = [r for r in rows if r["market"] == "GSS"]
    below = [r for r in gss if "lower" in r["ans"]["P 1"]]
    print(f"GSS pursuits: {len(gss)}, bidding below normal: {len(below)} (must be 0)")
    print(f"No-bid: {sum(1 for r in rows if r['bid']=='No Bid')} (must be 0)")
    print(f"Dependencies: {sum(1 for r in rows if r['dep'])}")
    print(f"No-bid: {sum(1 for r in rows if r['bid']=='No Bid')}")
    vals = sorted(r["value"] for r in rows)
    print(f"Total award value: ${sum(vals):,.0f}")
    print(f"Award value: median ${vals[len(vals)//2]:,.0f}  "
          f"min ${vals[0]:,.0f}  max ${vals[-1]:,.0f}")

    write_markets(args.markets, args.seed)
    write_xlsx(rows, args.xlsx)
    write_sql(rows, args.sql)
    json.dump([{k: (v.isoformat() if isinstance(v, datetime.date) else v)
                for k, v in r.items()} for r in rows],
              open(args.json, "w"), default=str)
    print(f"\nWrote {args.xlsx}, {args.sql}, {args.json}")


# ------------------------------------------------------------- markets
def write_markets(path, seed):
    """Two artifacts, deliberately separate.

    demo_markets.json         CDA-SIDE ONLY. Differentials. Never shipped to
                              a client database or workbook.
    demo_markets_public.json  What GET /v1/markets returns -- names only.
    demo_market_calibration.sql  Rows for the CDA-side calibration table.

    NOTE: ParseMarketsResponse_ in modPwinModelAdapter splits the response on
    commas, so a market NAME must never contain a comma."""
    # PINNED. These are already loaded in the AWS demo environment --
    # they must NOT change when the pipeline seed changes. MSN is neutral.
    DIFFS = {"GSS": -0.041, "ACP": 0.0007, "MSN": 0.000,
             "SUS": -0.018, "SPC": 0.025, "TRN": 0.027}
    recs = [dict(code=c, name=n, price_differential=DIFFS[c]) for c, n in MARKETS]

    json.dump({"client": "DEMO", "effective_from": f"{PY}-01-01",
               "calibration_version": "demo-1",
               "_warning": "CDA-SIDE ONLY. Price differentials are engine IP. "
                           "Never load into a client-facing database.",
               "markets": recs}, open(path, "w"), indent=2)

    pub = path.replace(".json", "_public.json")
    json.dump({"markets": [r["code"] for r in recs]}, open(pub, "w"), indent=2)

    sql = path.replace(".json", "_calibration.sql")
    L = ["-- CDA-SIDE ONLY. Do NOT load into the client-facing database.",
         "-- Price differentials are engine IP (decision 2b).",
         "-- market_calibration is keyed (client_code, market_code).", ""]
    for r in recs:
        L.append("INSERT INTO market_calibration "
                 "(client_code,market_code,price_differential,effective_from,calibration_version)")
        L.append(f"VALUES ('DEMO',{q(r['code'])},{r['price_differential']},"
                 f"DATE '{PY}-01-01','demo-1');")
    open(sql, "w").write("\n".join(L) + "\n")

    print("\nMarket differentials (CDA-side only):")
    for r in recs:
        flag = "  <- neutral" if r["price_differential"] == 0 else ""
        print(f"  {r['code']:5} {r['name']:32} {r['price_differential']:+7.2%}{flag}")


# ---------------------------------------------------------------- xlsx
AOP_COLS = ["UID", "Opportunity ID", "Market", "Dep", "Opportunity Name",
            "Planned Total Award Value", "Planned Fee", "Opportunity Type",
            "Sole Source", "Planned B&P Required", "B&P Start", "Proposal Due",
            "Contract Award", "End Date", "Bid/NoBid", "Phase",
            "Date Phase Completed", "SalesForce Pwin", "Won/Lost", "Cancel Date"]
PWIN_COLS = ["UID", "Opportunity Name", "Dependent UID", "Market", "Contract Type",
             "Bid Value", "Fee", "Investment", "Invest %", "Bidders",
             "TM 1a", "TM 1b", "TM 2", "TM 3", "TM 4", "TM 5", "PP 1", "P 1", "P 2"]
DEP_COLS = ["UID", "Opportunity Name", "Dependent UID", "Market", "Contract Type",
            "Base Pwin", "Fee", "Invest %", "Bidders",
            "TM 1a", "TM 1b", "TM 2", "TM 3", "TM 4", "TM 5", "PP 1", "P 1", "P 2"]


def write_xlsx(rows, path):
    import openpyxl
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def sheet(name, cols, data, table_name):
        ws = wb.create_sheet(name)
        ws.append(cols)
        for d in data:
            ws.append(d)
        ref = f"A1:{get_column_letter(len(cols))}{len(data)+1}"
        t = Table(displayName=table_name, ref=ref)
        t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(t)
        for i, c in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(i)].width = max(11, min(34, len(str(c)) + 4))
        for r in range(2, len(data) + 2):
            for i, c in enumerate(cols, 1):
                if "Date" in c or c in ("B&P Start", "Proposal Due", "Contract Award", "End Date"):
                    ws.cell(r, i).number_format = "yyyy-mm-dd"
                if c in ("Planned Total Award Value", "Planned B&P Required",
                         "Bid Value", "Investment"):
                    ws.cell(r, i).number_format = "#,##0"
                if c in ("Planned Fee", "Fee", "Invest %", "SalesForce Pwin", "Base Pwin"):
                    ws.cell(r, i).number_format = "0.00%"
        return ws

    aop = [[r["idx"], r["opp_id"], r["market"], r["dep"], r["name"], r["value"], r["fee"],
            r["otype"], r["sole"], r["bp"], r["bp_start"], r["due"], r["award"], r["end"],
            r["bid"], r["stage"],
            r["due"] if r["outcome"] in ("Won", "Lost") else None, None,
            r["outcome"] if r["outcome"] in ("Won", "Lost") else None,
            r["out_date"] if r["outcome"] == "Canceled" else None] for r in rows]
    sheet("Phased Opp, Budget, Rev Plan", AOP_COLS, aop, "AOP")

    pw = [[r["idx"], r["name"], r["dep"], r["market"], r["ctype"], r["value"], r["fee"],
           r["invest"], r["invest_pct"], r["bidders"]] +
          [r["ans"][k] for k in ["TM 1a", "TM 1b", "TM 2", "TM 3", "TM 4", "TM 5",
                                 "PP 1", "P 1", "P 2"]] for r in rows]
    sheet("Pursuits Pwins", PWIN_COLS, pw, "Pwin")

    # Dependent-scenario assessment: same pursuit, answered as if the
    # preceding pursuit was WON. Incumbency and price posture shift.
    rng = random.Random(99)
    dep = []
    for r in rows:
        if not r["dep"]:
            continue
        # Answered as if the preceding pursuit was WON: we are now incumbent.
        a = dict(r["ans"])
        if is_product(r["otype"]):
            if a["TM 1a"] == "Worse":
                a["TM 1a"] = "Same"
        else:
            a["TM 1a"] = "On contract today"
        a["TM 2"] = "Yes, us"
        a["TM 3"] = tm3_for("Yes, us", rng)
        tmp = dict(r, ans=a)
        enforce_gates(tmp)
        a = tmp["ans"]
        dep.append([r["idx"], r["name"], r["dep"], r["market"], r["ctype"], None,
                    r["fee"], r["invest_pct"], r["bidders"]] +
                   [a[k] for k in ["TM 1a", "TM 1b", "TM 2", "TM 3", "TM 4", "TM 5",
                                   "PP 1", "P 1", "P 2"]])
    sheet("Pursuits Pwins Dependent", DEP_COLS, dep, "Pwin_Dep")

    ws = wb.create_sheet("README", 0)
    for i, line in enumerate([
        ["CPDE synthetic demo pipeline"], [],
        ["Import via the CPDE workbook's Import button."],
        ["Only INPUT columns are supplied. Min/Max B&P, Calculated B&P, Planned"],
        ["Investment, Pwin, Probabilistic Award, all year phasing and all B&P"],
        ["Days columns are workbook formulas and rebuild on import."],
        ["Pwin scores and StaffingData are recalculated by the workbook."], [],
        ["All markets, names, values and dates are invented. No real data."], [],
        ["Pursuits", len(rows)],
        ["Planning year", PY],
        ["B&P oversubscription", f"{OVERSUBSCRIPTION:.0f}x budget, level loaded"],
    ], 1):
        ws.append(line)
    ws.column_dimensions["A"].width = 72
    wb.save(path)


# ---------------------------------------------------------------- sql
def q(v):
    if v is None:
        return "NULL"
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, (int, float)):
        return repr(v)
    if isinstance(v, datetime.date):
        return f"DATE '{v.isoformat()}'"
    return "'" + str(v).replace("'", "''") + "'"


OPP_CODE = {"Existing product solution, little modification required": "EXIST_PROD",
            "Developmental/New Product": "DEV_NEW",
            "Engineering/Technical Services": "ENG_TECH_SVC",
            "Sustainment/O&M": "SUSTAIN_OM"}
CTR_CODE = {"Cost Plus": "COST_PLUS", "Fixed Price": "FIXED_PRICE",
            "Time & Materials": "T_AND_M"}
STG_CODE = {"Pre-BH": "PRE_BH", "Post-BH": "POST_BH", "Post-PTW": "POST_PTW"}
OUT_CODE = {"Won": "WON", "Lost": "LOST", "Canceled": "CANCELLED"}


def write_sql(rows, path):
    L = ["-- CPDE synthetic demo pipeline. Generated by gen_demo.py.",
         "-- All markets, names, values and dates are invented.",
         "-- Assumes 01-04 schema + 03_seed reference data are loaded.",
         "BEGIN;", "",
         "INSERT INTO client (code,name) VALUES ('DEMO','Demo Aerospace')",
         "  ON CONFLICT (code) DO NOTHING;", "",
         "INSERT INTO org_node (client_id,parent_id,node_type,code,name)",
         "SELECT id,NULL,'business','BUS','Demo Aerospace' FROM client WHERE code='DEMO'",
         "  ON CONFLICT (client_id,code) DO NOTHING;",
         "INSERT INTO org_node (client_id,parent_id,node_type,code,name,is_license_boundary)",
         "SELECT c.id,o.id,'business_unit','BU1','Advanced Systems',TRUE",
         "  FROM client c JOIN org_node o ON o.client_id=c.id AND o.code='BUS'",
         " WHERE c.code='DEMO' ON CONFLICT (client_id,code) DO NOTHING;", ""]

    for i, (code, name) in enumerate(MARKETS, 1):
        L.append("INSERT INTO market (client_id,code,name,display_order) "
                 f"SELECT id,{q(code)},{q(name)},{i} FROM client WHERE code='DEMO' "
                 "ON CONFLICT (client_id,code) DO NOTHING;")
    L.append("")

    for y in YEARS:
        t = TARGETS[y]
        L.append("INSERT INTO plan_year (org_node_id,calendar_year,escalation_rate,"
                 "revenue_target,fee_target,budgeted_bp,budgeted_investment,"
                 "current_contract_revenue,current_contract_fee)")
        L.append(f"SELECT o.id,{y},{t['esc']},{t['rev']},{t['fee']},{t['bp']},{t['inv']},"
                 f"{CURRENT_BASE[y]},{round(CURRENT_BASE[y]*CURRENT_FEE_RATE)}")
        L.append("  FROM org_node o JOIN client c ON c.id=o.client_id")
        L.append(" WHERE c.code='DEMO' AND o.code='BU1' "
                 "ON CONFLICT (org_node_id,calendar_year) DO NOTHING;")
    L.append("")

    for r in rows:
        L.append(f"-- {r['idx']}: {r['name']}")
        L.append("""INSERT INTO pursuit (client_id,org_node_id,external_opportunity_id,name,
  market_id,opportunity_type_id,contract_type_id,pipeline_stage_id,
  is_sole_source,bidders,bid_decision,planned_total_award_value,planned_fee_rate,
  investment_pct,planned_investment,bp_start_date,proposal_due_date,
  contract_award_date,period_end_date,outcome,outcome_date)
SELECT c.id,o.id,%s,%s,m.id,ot.id,ct.id,ps.id,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s
  FROM client c
  JOIN org_node o ON o.client_id=c.id AND o.code='BU1'
  JOIN market m ON m.client_id=c.id AND m.code=%s
  JOIN opportunity_type ot ON ot.code=%s
  JOIN contract_type ct ON ct.code=%s
  JOIN pipeline_stage ps ON ps.code=%s
 WHERE c.code='DEMO';""" % (
            q(str(r["opp_id"])), q(r["name"]), q(r["sole"]), q(r["bidders"]),
            q("BID" if r["bid"] == "Bid" else "NO_BID"), q(r["value"]), q(r["fee"]),
            q(r["invest_pct"]), q(r["invest"]), q(r["bp_start"]), q(r["due"]),
            q(r["award"]), q(r["end"]),
            q(OUT_CODE.get(r["outcome"])), q(r["out_date"]),
            q(r["market"]), q(OPP_CODE[r["otype"]]), q(CTR_CODE[r["ctype"]]),
            q(STG_CODE[r["stage"]])))

    L.append("")
    L.append("-- dependencies (second pass: targets must exist first)")
    for r in rows:
        if r["dep"]:
            tgt = next(x for x in rows if x["idx"] == r["dep"])
            L.append("UPDATE pursuit p SET depends_on_pursuit_id=d.id "
                     "FROM pursuit d, client c "
                     f"WHERE c.code='DEMO' AND p.client_id=c.id AND d.client_id=c.id "
                     f"AND p.external_opportunity_id={q(str(r['opp_id']))} "
                     f"AND d.external_opportunity_id={q(str(tgt['opp_id']))};")

    L.append("")
    L.append("COMMIT;")
    L.append("")
    L.append("-- NOTE: pwin_assessment / pwin_answer rows are NOT generated here.")
    L.append("--       Pwin comes from the engine. Load answers and call /v1/pwin,")
    L.append("--       or import the xlsx into the workbook and let it recalculate.")
    open(path, "w").write("\n".join(L) + "\n")


if __name__ == "__main__":
    main()
